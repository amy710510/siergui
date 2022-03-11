import tkinter as tk
import tkinter.ttk as ttk
from tkinter import filedialog
from tkinter import messagebox
import pandas as pd
import openpyxl
from datetime import datetime

sier=tk.Tk() #建立主視窗，唯有'T'是大寫
sier.title('sier2021') #視窗標題
sier.geometry('1500x800')
sier.minsize(width=1500,height=800)

#分割視窗
div_size=100
f1=tk.Frame(sier,width=div_size*6, height=div_size)
f2=tk.Frame(sier,width=div_size*6, height=div_size*3)
f3=tk.Frame(sier,width=div_size*6, height=div_size*3)
f4=tk.Frame(sier,width=div_size*6, height=div_size)
f5=tk.Frame(sier,width=div_size, height=div_size)
f6=tk.Frame(sier,width=div_size, height=div_size)

f1.grid(row=0, column=0, sticky="news") #sticky元件對齊方式，news上下左右都可以伸展
f2.grid(row=1, column=0, sticky="news")
f3.grid(row=2, column=0, sticky="news")
f4.grid(row=3, column=0, sticky="news")
f5.grid(row=0, column=1, sticky="news",rowspan=3)
f6.grid(row=3, column=1, sticky="news")

#隨視窗大小調整縮放比例，weight是每一個row與column放大的權重
sier.columnconfigure(0,weight=1)
sier.columnconfigure(1,weight=1)
sier.rowconfigure(0,weight=1)
sier.rowconfigure(1,weight=1)
sier.rowconfigure(2,weight=1)
sier.rowconfigure(3,weight=1)

#匯入與讀取檔案
op_path=''  #宣告全域變數，在def函式以外也可使用https://shengyu7697.github.io/python-global/
#global op_path：改變全域變數時，須先加入此行
#header：用row=0作列名
def op():
    global op_path
    op_path=filedialog.askopenfilename(title='選擇',filetypes=[('Excel',('*.xls','*.xlsx'))])
    rawdata=pd.read_excel(op_path,header=0)
    return rawdata

data=op()

#設定檔名
#defaultextension副檔名
#openpyxl.load_workbook()讀取excel工作檔https://hackmd.io/@amostsai/SJkC1_EcX?type=view
df=''
writer=''
writerfile=''
def export():
    global df
    global writer
    global writerfile
    df=pd.DataFrame(columns=["文本ID", "事件ID",'文本紀錄','事件紀錄','備註',"朝代",'起始季節','中曆起始年','中曆起始月',
                             '中曆起始日','迄止季節','中曆迄止年','中曆迄止月','中曆迄止日','西曆起始年','西曆起始月','西曆起始日',
                             '西曆迄止年','西曆迄止月','西曆迄止日','古地名','今隸屬省級','今隸屬縣市','經度','緯度','高度',
                             '縣市ID','事件編碼','傷亡人數','受災戶數','文獻','書名','冊','頁','錯誤與疑問','錯誤代碼',
                             '同義詞','登錄人代碼','登錄日期','完成度代碼','登錄-資料庫版本','登錄-編碼版本','登錄-軟體版本'])
    savefile=filedialog.asksaveasfilename(defaultextension=".xlsx",title='選擇',filetypes=[('Excel','*.xlsx')])
    writer=pd.ExcelWriter(savefile,engine='openpyxl')
    df.to_excel(writer,sheet_name='Sheet')
    writer.save()
    writerfile=openpyxl.load_workbook(writer)
    sheet=writerfile.active
        
exportfile=tk.Button(f6,text='設定檔名',command=export)
exportfile.grid(row=1,column=0,pady=5)#y軸間距5



#資料選單
datainfo6=tk.Label(f5,text='資料選單')
datainfo6.grid(row=0,column=0)

#抽出id並轉成串列
list_id=list(data.iloc[:,0])

datascrollbar=tk.Scrollbar(sier)
datascrollbar.grid(row=0,column=1,rowspan=3,pady=30,sticky='nse')
datalist=tk.Listbox(sier,width=30,yscrollcommand=datascrollbar.set) #此為垂直捲軸，水平捲軸=xscrollcommand

#將串列經enumerate()變成索引序列
for x,y in enumerate(list_id):
    datalist.insert(x,y)

datalist.grid(row=0,column=1,rowspan=3,pady=30,sticky='nsw')
datascrollbar.config(command=datalist.yview)


#匯入資料
#  .curselection()：取得項目索引
#  .delete(0,'end')：刪除欄位中第0個字元，到第end字元，等同於刪除欄位中的所有內容。
# tk.Text.delete('1.0','end')：需要以符點數指定字元
# label寫法：id1en.config(text=data.iloc[value,0]) ，volla.config(text=data.iloc[value,14])
def update():
    value=datalist.curselection()[0]
    id1en.delete(0,'end')
    id1en.insert(0,data.iloc[value,0])
    id2en.delete(0,'end')
    idtext=id1en.get()[9:11]+'-000'
    id2en.insert(0,idtext)
    volen.delete(0,'end')
    volen.insert(0,data.iloc[value,14])
    pageen.delete(0,'end')
    pageen.insert(0,data.iloc[value,17])
    
    dynastyen.delete(0,'end')
    dynastyen.insert(0,data.iloc[value,1])
    year_len.delete(0,'end')
    year_len.insert(0,data.iloc[value,2])
    month_len.delete(0,'end')
    day_len.delete(0,'end')
    year_len2.delete(0,'end')
    year_len2.insert(0,data.iloc[value,3])
    month_len2.delete(0,'end')
    day_len2.delete(0,'end')
    sea_1en.delete(0,'end')
    year_wen.delete(0,'end')
    year_wen.insert(0,int(data.iloc[value,4]))
    month_wen.delete(0,'end')
    day_wen.delete(0,'end')
    sea_2en.delete(0,'end')
    year_wen2.delete(0,'end')
    year_wen2.insert(0,data.iloc[value,5])
    month_wen2.delete(0,'end')
    day_wen2.delete(0,'end')
     
    geo1en.delete(0,"end")
    geo1en.insert(1,data.iloc[value,7])
    geo2en.delete(0,"end")
    geo2en.insert(1,data.iloc[value,6])
    geo3en.delete(0,"end")
    geo3en.insert(1,data.iloc[value,9])
    datafr1te.delete('1.0',"end")
    datafr1te.insert('1.0',data.iloc[value,12])
    datafr2en.delete(0,"end")
    datafr2en.insert(1,data.iloc[value,13][0:13])
    datafr3en.delete(0,"end")
    datafr4en.delete(0,"end")
    datafr7en.delete(0,"end")
    datafr8en.delete(0,"end")
    datafr9en.delete(0,"end")
    
    data1te.delete('1.0',"end")
    data1te.insert('1.0',data.iloc[value,10])
    data2te.delete('1.0',"end")
    data3te.delete('1.0',"end")
    data3te.insert('1.0',data.iloc[value,11])
    disasteren.delete(0,"end")
    disasteren1.delete(0,"end")
    return None
    
displaydata=tk.Button(f6,text='匯入資料',command=update)
displaydata.grid(row=2,column=0,pady=5)
    

#label與輸入框
#每筆資料資訊-第一列
datainfo1=tk.Label(f1,text='資料編號')
datainfo1.grid(row=0,column=0)

vol=tk.Label(f1,text='冊：')
vol.grid(row=0,column=1,sticky='w')
volen=tk.Entry(f1)
volen.grid(row=0,column=2)
id1=tk.Label(f1,text='文本ID：')
id1.grid(row=0,column=4)
id1en=tk.Entry(f1,text='')
id1en.grid(row=0,column=5)
id2=tk.Label(f1,text='事件ID：')
id2.grid(row=0,column=6)
id2en=tk.Entry(f1)
id2en.grid(row=0,column=7)
page=tk.Label(f1,text='頁碼：')
page.grid(row=0,column=8)
pageen=tk.Entry(f1)
pageen.grid(row=0,column=9)

#每筆資料-文本時間
datainfo3=tk.Label(f1,text='時間')
datainfo3.grid(row=1,column=0,sticky='w')

dynasty=tk.Label(f1,text='朝代：')
dynasty.grid(row=1,column=1,sticky='w')
dynastyen=tk.Entry(f1)
dynastyen.grid(row=1,column=2,sticky='w')
year_l=tk.Label(f1,text='中曆起始年：')
year_l.grid(row=1,column=4,sticky='w')
year_len=tk.Entry(f1)
year_len.grid(row=1,column=5)
month_l=tk.Label(f1,text='中曆起始月：')
month_l.grid(row=1,column=6,sticky='w')
month_len=tk.Entry(f1)
month_len.grid(row=1,column=7)
day_l=tk.Label(f1,text='中曆起始日：')
day_l.grid(row=1,column=8,sticky='w')
day_len=tk.Entry(f1)
day_len.grid(row=1,column=9)
year_l2=tk.Label(f1,text='中曆迄止年：')
year_l2.grid(row=2,column=4,sticky='w')
year_len2=tk.Entry(f1)
year_len2.grid(row=2,column=5)
month_l2=tk.Label(f1,text='中曆迄止月：')
month_l2.grid(row=2,column=6,sticky='w')
month_len2=tk.Entry(f1)
month_len2.grid(row=2,column=7)
day_l2=tk.Label(f1,text='中曆迄止日：')
day_l2.grid(row=2,column=8,sticky='w')
day_len2=tk.Entry(f1)
day_len2.grid(row=2,column=9)
sea_1=tk.Label(f1,text='起始季節：')
sea_1.grid(row=3,column=1,sticky='w')
sea_1en=tk.Entry(f1)
sea_1en.grid(row=3,column=2)
year_w=tk.Label(f1,text='西曆起始年：')
year_w.grid(row=3,column=4,sticky='w')
year_wen=tk.Entry(f1)
year_wen.grid(row=3,column=5)
month_w=tk.Label(f1,text='西曆起始月：')
month_w.grid(row=3,column=6,sticky='w')
month_wen=tk.Entry(f1)
month_wen.grid(row=3,column=7)
day_w=tk.Label(f1,text='西曆起始日：')
day_w.grid(row=3,column=8,sticky='w')
day_wen=tk.Entry(f1)
day_wen.grid(row=3,column=9)
sea_2=tk.Label(f1,text='迄止季節：')
sea_2.grid(row=4,column=1,sticky='w')
sea_2en=tk.Entry(f1)
sea_2en.grid(row=4,column=2)
year_w2=tk.Label(f1,text='西曆迄止年：')
year_w2.grid(row=4,column=4,sticky='w')
year_wen2=tk.Entry(f1)
year_wen2.grid(row=4,column=5)
month_w2=tk.Label(f1,text='西曆迄止月：')
month_w2.grid(row=4,column=6,sticky='w')
month_wen2=tk.Entry(f1)
month_wen2.grid(row=4,column=7)
day_w2=tk.Label(f1,text='西曆迄止日：')
day_w2.grid(row=4,column=8,sticky='w')
day_wen2=tk.Entry(f1)
day_wen2.grid(row=4,column=9)


#季節轉換西曆日期
#防呆機制：季節轉換西曆日期時，藉計算時間差，檢查季節年份是否有誤
def seasonstart():
    global year_wen
    global month_wen
    global day_wen
    global year_wen2
    global month_wen2
    global day_wen2
    month_wen.delete(0,'end')
    day_wen.delete(0,'end')
    month_wen2.delete(0,'end')
    day_wen2.delete(0,'end')
    if (sea_1en.get()=='春'):
        month_wen.insert(0,2)
        day_wen.insert(0,3)
    elif (sea_1en.get()=='夏'):
        month_wen.insert(0,5)
        day_wen.insert(0,5)
    elif (sea_1en.get()=='秋'):
        month_wen.insert(0,8)
        day_wen.insert(0,7)            
    elif (sea_1en.get()=='冬'):
        month_wen.insert(0,11)
        day_wen.insert(0,7)
    else :
        return None

    def seasonend():
        if (sea_2en.get()=='春'):
            month_wen2.insert(0,5)
            day_wen2.insert(0,4)
        elif (sea_2en.get()=='夏'):
            month_wen2.insert(0,8)
            day_wen2.insert(0,6)
        elif (sea_2en.get()=='秋'):
            month_wen2.insert(0,11)
            day_wen2.insert(0,6)
        elif (sea_2en.get()=='冬'):
            month_wen2.insert(0,2)
            day_wen2.insert(0,2)
            text2=int(year_wen2.get())+1
            year_wen2.delete(0,'end')
            year_wen2.insert(0,text2)                
        else :
            return None
    def days():
        startdate=datetime(int(year_wen.get()),int(month_wen.get()),int(day_wen.get()))
        enddate=datetime(int(year_wen2.get()),int(month_wen2.get()),int(day_wen2.get()))
        thedays=enddate-startdate
        if (thedays.days<0):
            warning=messagebox.showwarning(title="有誤",message='西曆迄止日期早於起始，請再次確認')
        else:
            return None
    
    seasonend()
    days()

seasontodate=tk.Button(f1,text='季節轉換',command=seasonstart)
seasontodate.grid(row=4,column=0)     


#每筆資料-文本內容
datainfo2=tk.Label(f2,text='內文')
datainfo2.grid(row=0,column=0,sticky='nw')

data1=tk.Label(f2,text='文本紀錄：')
data1.grid(row=0,column=1,sticky='nw')
data1te=tk.Text(f2,width=150,height=5)
data1te.grid(row=0,column=2,sticky='nw')
data2=tk.Label(f2,text='事件紀錄：')
data2.grid(row=1,column=1,sticky='nw',pady=10)
data2te=tk.Text(f2,width=150,height=5)
data2te.grid(row=1,column=2,pady=10,sticky='nw')
data3=tk.Label(f2,text='備註：')
data3.grid(row=2,column=1,sticky='nw',pady=5)
data3te=tk.Text(f2,width=150,height=3)
data3te.grid(row=2,column=2,pady=10,sticky='nw')

#疫災事件編碼
disaster=tk.Label(f3,text='疫災編碼')
disaster.grid(row=0,column=0,sticky='w')
disaster1=tk.Label(f3,text='主類別分類')
disaster1.grid(row=0,column=1,sticky='w')
disaster11=tk.Label(f3,text='未判定：-98，無敘述/不處理：-9999')
disaster11.grid(row=0,column=3,sticky='w')
disaster2=tk.Label(f3,text='主類別')
disaster2.grid(row=1,column=1,sticky='w',pady=2)
disaster3=tk.Label(f3,text='次類別')
disaster3.grid(row=2,column=1,sticky='w',pady=2)
disaster4=tk.Label(f3,text='詞彙')
disaster4.grid(row=3,column=1,sticky='w',pady=2)
disaster5=tk.Label(f3,text='程度')
disaster5.grid(row=4,column=1,sticky='w',pady=2)
disaster6=tk.Label(f3,text='時間')
disaster6.grid(row=5,column=1,sticky='w',pady=2)

a_dict={}
a_list=[]
b_dict={}
b_list=[]
c_dict={}
c_list=[]
d_dict={}
d_list=[]
e_dict={}
e_list=[]
f_dict={}
f_list=[]

wb=openpyxl.load_workbook('氣象疫災編碼.xlsx')
ws1=wb['MasterEvent']
ws2=wb['MasterEventClass']
ws3=wb['ServantEvent']
ws4=wb['ServantEventClass']
ws5=wb['LevelEventClass']
ws6=wb['TimeEventClass']

#num1na、num1id讀取主類別分類的名稱與數字
#for x in range(1,ws1.max_row+1)：row+1才能讀取到最後一行資料
#a_dict.update()主類別分類名稱與數字保存在字典
#a_list.append()主類別分類名稱保存在list
for x in range(1,ws1.max_row+1):
    num1na=ws1.cell(row=x,column=1).value
    num1id=ws1.cell(row=x,column=2).value
    a_dict.update({num1na:num1id})
    a_list.append(num1na)

num1=tk.StringVar()
dcombo1=ttk.Combobox(f3,textvariable=num1)
dcombo1.grid(row=0,column=2,ipadx=250,sticky='w')
dcombo1['value']=(a_list)

num2=tk.StringVar()
dcombo2=ttk.Combobox(f3,textvariable=num2)
dcombo2.grid(row=1,column=2,ipadx=250,sticky='w')

num3=tk.StringVar()
dcombo3=ttk.Combobox(f3,textvariable=num3)
dcombo3.grid(row=2,column=2,ipadx=250,sticky='w')

num4=tk.StringVar()
dcombo4=ttk.Combobox(f3,textvariable=num4)
dcombo4.grid(row=3,column=2,ipadx=250,sticky='w')

num5=tk.StringVar()
dcombo5=ttk.Combobox(f3,textvariable=num5)
dcombo5.grid(row=4,column=2,ipadx=250,sticky='w')

num6=tk.StringVar()
dcombo6=ttk.Combobox(f3,textvariable=num6)
dcombo6.grid(row=5,column=2,ipadx=250,sticky='w')

def afunc(event):
    a_dict_str=a_dict[num1.get()]
    b_dict.clear()
    b_list.clear()
    for xx in range(1,ws2.max_row+1):
        num2na=ws2.cell(row=xx,column=2).value
        num2id=ws2.cell(row=xx,column=3).value
        num2id_str=ws2.cell(row=xx,column=1).value
        if num2id_str==a_dict_str:
            b_dict.update({num2na:num2id})
            b_list.append(num2na)
    dcombo2['value']=(b_list)

def bfunc(event):
    b_dict_str=b_dict[num2.get()]
    c_dict.clear()
    c_list.clear()
    for xxx in range(1,ws3.max_row+1):
        num3na=ws3.cell(row=xxx,column=2).value
        num3id=ws3.cell(row=xxx,column=3).value
        num3id_str=ws3.cell(row=xxx,column=1).value
        if num3id_str==b_dict_str:
            c_dict.update({num3na:num3id})
            c_list.append(num3na)
    dcombo3['value']=(c_list)

def cfunc(event):
    c_dict_str=int(str(b_dict[num2.get()])+str(c_dict[num3.get()]))
    d_dict.clear()
    d_list.clear()
    for xxxx in range(1,ws4.max_row+1):
        num4na=ws4.cell(row=xxxx,column=2).value
        num4id=ws4.cell(row=xxxx,column=3).value
        num4id_str=ws4.cell(row=xxxx,column=1).value
        num45id=ws4.cell(row=xxxx,column=4).value
        if num4id_str==c_dict_str:
            d_dict.update({num4na:num4id})
            d_list.append(num4na)
    dcombo4['value']=(d_list)
    
def dfunc(event):
    b_dict_str=b_dict[num2.get()]
    e_dict.clear()
    e_list.clear()
    for xxxxx in range(1,ws5.max_row+1):
        num5na=ws5.cell(row=xxxxx,column=2).value
        num5id=ws5.cell(row=xxxxx,column=3).value
        num5id_str=ws5.cell(row=xxxxx,column=1).value
        if num5id_str==b_dict_str:
            e_dict.update({num5na:num5id})
            e_list.append(num5na)
    dcombo5['value']=(e_list)

def efunc(event):
    b_dict_str=b_dict[num2.get()]
    f_dict.clear()
    f_list.clear()
    for xxxxxx in range(1,ws6.max_row+1):
        num6na=ws6.cell(row=xxxxxx,column=2).value
        num6id=ws6.cell(row=xxxxxx,column=3).value
        num6id_str=ws6.cell(row=xxxxxx,column=1).value
        if num6id_str==b_dict_str:
            f_dict.update({num6na:num6id})
            f_list.append(num6na)
    dcombo6['value']=(f_list)

def ffunc(event):
    disastertext.set(str(b_dict[dcombo2.get()])+str(c_dict[dcombo3.get()])+str(d_dict[num4.get()])
                     +str(e_dict[dcombo5.get()])+str(f_dict[dcombo6.get()])
                     +dcombo2.get()+str(' - ')+dcombo3.get()+str(' - ')+num4.get()+str(' - ')+dcombo5.get()+str(' - ')
                     +dcombo6.get())
    
dcombo1.bind('<<ComboboxSelected>>',afunc)
dcombo2.bind('<<ComboboxSelected>>',bfunc)
dcombo3.bind('<<ComboboxSelected>>',cfunc)  
dcombo4.bind('<<ComboboxSelected>>',dfunc)  
dcombo5.bind('<<ComboboxSelected>>',efunc)
dcombo6.bind('<<ComboboxSelected>>',ffunc) 

disastertext=tk.StringVar()
disasterla=tk.Label(f3,text='編碼')
disasterla.grid(row=6,column=1,sticky='w')
disasteren=tk.Entry(f3,textvariable=disastertext) #textvariable=disastertext：stringvar的值傳至此輸入框
disasteren.grid(row=6,column=2,ipadx=300,sticky='w')

def disasterid():
    if disasteren1.get()=='':
        text3=disasteren.get()[0:9]
        disasteren1.insert(0,text3)
    else:
        text4=';'+disasteren.get()[0:9]
        disasteren1.insert('end',text4)

disasterbu=tk.Button(f3,text='新增編碼',command=disasterid)
disasterbu.grid(row=7,column=1)
disasteren1=tk.Entry(f3) 
disasteren1.grid(row=7,column=2,ipadx=300,sticky='w')


#每筆資料-地名空間與文獻
datainfo4=tk.Label(f4,text='空間')
datainfo4.grid(row=0,column=0)

geo1=tk.Label(f4,text='古地名：')
geo1.grid(row=0,column=1,sticky='w')
geo1en=tk.Entry(f4)
geo1en.grid(row=0,column=2,sticky='w')
geo2=tk.Label(f4,text='今隸屬省級：')
geo2.grid(row=0,column=3,sticky='w')
geo2en=tk.Entry(f4)
geo2en.grid(row=0,column=4)
geo3=tk.Label(f4,text='今隸屬縣市：')
geo3.grid(row=0,column=5,sticky='w')
geo3en=tk.Entry(f4)
geo3en.grid(row=0,column=6)

datafr1=tk.Label(f4,text='文獻：')
datafr1.grid(row=1,column=1,sticky='w')
datafr1te=tk.Text(f4,width=59,height=4)
datafr1te.grid(row=1,column=2,pady=2,sticky='w')
datafr2=tk.Label(f4,text='書名：')
datafr2.grid(row=2,column=1,sticky='w')
datafr2en=tk.Entry(f4,width=50)
datafr2en.grid(row=2,column=2,ipadx=30)
datafr3=tk.Label(f4,text='錯誤與疑問：')
datafr3.grid(row=3,column=1,sticky='w')
datafr3en=tk.Entry(f4,width=50)
datafr3en.grid(row=3,column=2,ipadx=30)
datafr4=tk.Label(f4,text='同義詞：')
datafr4.grid(row=4,column=1,sticky='w')
datafr4en=tk.Entry(f4,width=50)
datafr4en.grid(row=4,column=2,ipadx=30)
datafr5=tk.Label(f4,text='登錄人代碼：')
datafr5.grid(row=5,column=1,sticky='w')
datafr5en=tk.Entry(f4)
datafr5en.grid(row=5,column=2,sticky='w')
datafr6=tk.Label(f4,text='登錄日期：')
datafr6.grid(row=5,column=3,sticky='w')
datafr6en=tk.Entry(f4)
datafr6en.grid(row=5,column=4)
datafr7=tk.Label(f4,text='完成度代碼：')
datafr7.grid(row=5,column=5,sticky='w')
datafr7en=tk.Entry(f4)
datafr7en.grid(row=5,column=6)
datafr8=tk.Label(f4,text='傷亡人數：')
datafr8.grid(row=3,column=3,sticky='w')
datafr8en=tk.Entry(f4)
datafr8en.grid(row=3,column=4,sticky='w')
datafr9=tk.Label(f4,text='受災戶數：')
datafr9.grid(row=3,column=5,sticky='w')
datafr9en=tk.Entry(f4)
datafr9en.grid(row=3,column=6,sticky='w')

#逐筆匯出資料
#sheet.max_row：工作表的總row數
#防呆機制：藉計算時間差，檢查手動輸入的西曆日期是否有誤。有誤時，無法匯出此筆資料。
def newfile():
    if (month_wen.get()!='' and day_wen.get()!='' and month_wen2.get()!='' and day_wen2.get()!=''):
        startdate=datetime(int(year_wen.get()),int(month_wen.get()),int(day_wen.get()))
        enddate=datetime(int(year_wen2.get()),int(month_wen2.get()),int(day_wen2.get()))
        thedays=enddate-startdate
        while (thedays.days<0):
            warning=messagebox.showwarning(title="有誤",message='西曆迄止日期早於起始，請再次確認')
            return
    else:
        print("系統未檢查西曆日期")

    sheet=writerfile.active
    sheet.cell(row=sheet.max_row+1,column=2).value='sier_disaster_9787533337957_'+id1en.get()
    sheet.cell(row=sheet.max_row,column=3).value=id2en.get()
    sheet.cell(row=sheet.max_row,column=4).value=data1te.get('1.0',"end")
    sheet.cell(row=sheet.max_row,column=5).value=data2te.get('1.0',"end")
    sheet.cell(row=sheet.max_row,column=6).value=data3te.get('1.0',"end")
        
    sheet.cell(row=sheet.max_row,column=7).value=dynastyen.get()
    sheet.cell(row=sheet.max_row,column=8).value=sea_1en.get()
    sheet.cell(row=sheet.max_row,column=9).value=year_len.get()
    sheet.cell(row=sheet.max_row,column=10).value=month_len.get()
    sheet.cell(row=sheet.max_row,column=11).value=day_len.get()
    sheet.cell(row=sheet.max_row,column=12).value=sea_2en.get()
    sheet.cell(row=sheet.max_row,column=13).value=year_len2.get()
    sheet.cell(row=sheet.max_row,column=14).value=month_len2.get()
    sheet.cell(row=sheet.max_row,column=15).value=day_len2.get()
    sheet.cell(row=sheet.max_row,column=16).value=year_wen.get()
    sheet.cell(row=sheet.max_row,column=17).value=month_wen.get()
    sheet.cell(row=sheet.max_row,column=18).value=day_wen.get()
    sheet.cell(row=sheet.max_row,column=19).value=year_wen2.get()
    sheet.cell(row=sheet.max_row,column=20).value=month_wen2.get()
    sheet.cell(row=sheet.max_row,column=21).value=day_wen2.get()
        
    sheet.cell(row=sheet.max_row,column=22).value=geo1en.get()
    sheet.cell(row=sheet.max_row,column=23).value=geo2en.get()
    sheet.cell(row=sheet.max_row,column=24).value=geo3en.get()
        
    sheet.cell(row=sheet.max_row,column=29).value=disasteren1.get()
    sheet.cell(row=sheet.max_row,column=30).value=datafr8en.get()
    sheet.cell(row=sheet.max_row,column=31).value=datafr9en.get()
    sheet.cell(row=sheet.max_row,column=32).value=datafr1te.get('1.0',"end")
    sheet.cell(row=sheet.max_row,column=33).value=datafr2en.get()
    voldict={'Ⅰ':'v.1先秦至明代卷','II':'v.2清代卷','V':'v.5畜疫卷'}
    sheet.cell(row=sheet.max_row,column=34).value=voldict[volen.get()]
    sheet.cell(row=sheet.max_row,column=35).value=pageen.get()
    sheet.cell(row=sheet.max_row,column=36).value=datafr3en.get() 
    sheet.cell(row=sheet.max_row,column=38).value=datafr4en.get()
    sheet.cell(row=sheet.max_row,column=39).value=datafr5en.get() 
    sheet.cell(row=sheet.max_row,column=40).value=datafr6en.get() 
    sheet.cell(row=sheet.max_row,column=41).value=datafr7en.get()
    sheet.cell(row=sheet.max_row,column=42).value=str('v1.1')
    sheet.cell(row=sheet.max_row,column=43).value=str('v4.6')
    sheet.cell(row=sheet.max_row,column=44).value=str('v1.0.3')
    writerfile.save(writer)
    print(id1en.get(),"儲存成功")    

getfile=tk.Button(f6,text='儲存檔案',command=newfile)
getfile.grid(row=3,column=0,pady=5)#y軸間距5

#主視窗需常駐在畫面上，非短暫出現
sier.mainloop() 
